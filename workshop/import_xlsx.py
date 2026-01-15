"""Import data from the Excel spreadsheet into SQLite."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

from workshop.db import WorkshopDB


def import_from_xlsx(xlsx_path: Path | str, db: WorkshopDB) -> None:
    """Import all data from the workshop Excel spreadsheet."""
    wb = openpyxl.load_workbook(xlsx_path)

    # Import organizers
    print("Importing organizers...")
    sheet = wb["Organization Committee"]
    conn = db._get_conn()
    conn.execute("DELETE FROM organizers")  # Clear existing
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            conn.execute(
                "INSERT INTO organizers (name, affiliation, email, email_secondary) VALUES (?, ?, ?, ?)",
                (row[0], row[1], row[2], row[3] if len(row) > 3 else None),
            )
    conn.commit()

    # Import guests
    print("Importing guests...")
    sheet = wb["Guest List"]
    conn.execute("DELETE FROM guests")  # Clear existing
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            conn.execute(
                """INSERT INTO guests (name, affiliation, status, type, notes, tags,
                   email, proposed_by, area, stipend, honorarium)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    row[0],  # name
                    row[1],  # affiliation
                    row[2] or "Proposed",  # status
                    row[3],  # type
                    row[4],  # notes
                    row[5],  # tags
                    row[6],  # email
                    row[7],  # proposed_by
                    row[8],  # area
                    float(row[9]) if row[9] else 0,  # stipend
                    float(row[10]) if row[10] else 0,  # honorarium
                ),
            )
    conn.commit()

    # Import budget
    print("Importing budget...")
    sheet = wb["Budget"]
    conn.execute("DELETE FROM budget")
    for row in sheet.iter_rows(min_row=2, max_row=9, values_only=True):
        if row[0] and row[1] and isinstance(row[1], (int, float)):
            conn.execute(
                "INSERT INTO budget (item, amount, notes) VALUES (?, ?, ?)",
                (row[0], float(row[1]), row[2]),
            )
    conn.commit()

    # Import travel costs
    print("Importing travel costs...")
    sheet = wb["Travel Costs"]
    conn.execute("DELETE FROM travel_costs")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            conn.execute(
                "INSERT INTO travel_costs (origin, avg_fare) VALUES (?, ?)",
                (row[0], float(row[1])),
            )
    conn.commit()

    # Import topics
    print("Importing topics...")
    sheet = wb["TopicsFormat Recommendations"]
    conn.execute("DELETE FROM topics")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            conn.execute(
                "INSERT INTO topics (event, type, suggested_by, notes) VALUES (?, ?, ?, ?)",
                (row[0], row[1], row[2], row[3]),
            )
    conn.commit()

    # Import schedule
    print("Importing schedule...")
    sheet = wb["Schedule"]
    conn.execute("DELETE FROM schedule")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        day = None
        if row[0]:
            if isinstance(row[0], datetime):
                day = row[0].date().isoformat()
            else:
                day = str(row[0])

        start_time = None
        if row[1]:
            if hasattr(row[1], "isoformat"):
                start_time = row[1].isoformat()
            else:
                start_time = str(row[1])

        end_time = None
        if row[2]:
            if hasattr(row[2], "isoformat"):
                end_time = row[2].isoformat()
            else:
                end_time = str(row[2])

        if day or row[3]:  # Has day or event
            conn.execute(
                "INSERT INTO schedule (day, start_time, end_time, event) VALUES (?, ?, ?, ?)",
                (day, start_time, end_time, row[3]),
            )
    conn.commit()
    conn.close()

    print("Import complete!")


if __name__ == "__main__":
    import sys

    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else (
        Path(__file__).parent.parent
        / "Workshop on AI & Indigenous Language Revitalization"
        / "Workshop on AI and Indigenous Languages Data.xlsx"
    )

    db = WorkshopDB()
    import_from_xlsx(xlsx_path, db)
    print(f"Database created at: {db.db_path}")
